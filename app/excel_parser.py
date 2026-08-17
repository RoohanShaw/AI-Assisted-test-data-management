"""
excel_parser.py — Deterministic extraction layer. ZERO AI.

Reads either:
  • SampleInput.json  — the structured JSON produced by the test management tool
  • TP_*.xlsx         — the Excel TDM template file

Excel Format (TP_AppointmentList_TDM.xlsx):
  - SheetMap sheet:  maps ExcelSheet names → TestSuite names
  - Login Page:      TestSuiteName | ObjectName | Url | UserId | Password | ...
  - Sheet1..N:       TS_XX | ... | SheetMap
                     ObjectName | Iteration Value | OCRTextVerification | Field Names | f1 | f2 ...
                     <ObjectName> | | | | | ...
                     | 1 | | | | ...   (iteration rows — one per iteration)
  - Lookup:          TestSuiteName | Object | ParentField | Iteration | Field Names | ...

Outputs a Normalized Internal JSON (ParsedTemplate) with all test suites,
objects, iterations, fields, login placeholders, and lookup placeholders.

Field metadata (Type, IsNavigation, IsLookupon, IsMandatory) is preserved in
'field_metadata' dict per object so the pipeline can perform deterministic
type-based routing before invoking the AI (FAISS) pipeline.

AI starts AFTER this module. This module does only deterministic, structural work.
"""

import json
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────
# Field-type constants (from the test tool schema)
# ──────────────────────────────────────────────
# Type 24 = button / navigation action (skip for data generation)
NAV_TYPE = 24

# Standard login keys always present in the output
LOGIN_KEYS = ["url", "userid", "password", "mobile no", "branch name", "dashboard url"]


# ──────────────────────────────────────────────
# Data structures (plain dicts — no Pydantic here)
# ──────────────────────────────────────────────

def _empty_object_node() -> Dict[str, Any]:
    """Skeleton for one Selenium flow object."""
    return {
        "iteration_count": 1,
        "fields": [],            # list of display field names for data generation
        "lookup_fields": [],     # fields with IsLookupon=true
        "has_login": True,       # login block always present
        "field_metadata": {},    # field_name → {type, is_navigation, is_lookup, is_mandatory}
    }


# ──────────────────────────────────────────────
# JSON Input Parser (primary path)
# ──────────────────────────────────────────────

def parse_json_input(data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Parse a SampleInput.json dict into a normalized internal structure.

    Args:
        data: Parsed JSON dict from SampleInput.json

    Returns:
        Normalized template dict of the form:
        {
          "test_package_name": str,
          "test_suites": {
            "TS_01": {
              "objects": {
                "Object Name": {
                  "iteration_count": int,
                  "fields": [str, ...],
                  "lookup_fields": [str, ...],
                  "has_login": True
                }
              }
            }
          }
        }
    """
    package_name: str = data.get("PackageName", "UnknownPackage")
    test_suites_raw: List[Dict] = data.get("TestSuites", [])

    normalized: Dict[str, Any] = {
        "test_package_name": package_name,
        "test_suites": {},
    }

    for suite in test_suites_raw:
        suite_name: str = suite.get("TestSuiteName", "TS_UNKNOWN")
        selected_suites: List[Dict] = suite.get("SelectedTestSuite", [])

        suite_node: Dict[str, Any] = {"objects": {}}

        for scenario in selected_suites:
            iteration_count: int = _safe_int(scenario.get("IterationCount", 1), default=1)
            flows: List[Dict] = scenario.get("SeleniumExecutionFlow", [])

            for flow in flows:
                obj_name: str = flow.get("Name", "").strip()
                if not obj_name:
                    continue

                # Deduplicate: if same object name already exists, merge fields
                if obj_name not in suite_node["objects"]:
                    suite_node["objects"][obj_name] = _empty_object_node()

                obj = suite_node["objects"][obj_name]

                # Use the max iteration count seen for this object
                obj["iteration_count"] = max(obj["iteration_count"], iteration_count)

                # Extract data fields — skip navigation-only buttons (Type=24)
                fields_raw: List[Dict] = flow.get("Fields", [])
                data_fields, lookup_fields, field_meta = _extract_fields(fields_raw)

                # Prefer DisplayName[0] when available and non-empty, else Name
                for field_name in data_fields:
                    if field_name not in obj["fields"]:
                        obj["fields"].append(field_name)

                for field_name in lookup_fields:
                    if field_name not in obj["lookup_fields"]:
                        obj["lookup_fields"].append(field_name)

                # Merge field metadata (first-seen wins for duplicates)
                for field_name, meta in field_meta.items():
                    if field_name not in obj["field_metadata"]:
                        obj["field_metadata"][field_name] = meta

        normalized["test_suites"][suite_name] = suite_node

    logger.info(
        f"[excel_parser] Parsed JSON: package='{package_name}', "
        f"suites={list(normalized['test_suites'].keys())}"
    )
    return normalized


def _extract_fields(
    fields_raw: List[Dict],
) -> tuple[List[str], List[str], Dict[str, Dict]]:
    """
    Separate data fields from lookup fields. Skip pure navigation/button types.
    Also returns field_metadata: a dict mapping field_name → {type, is_navigation, is_lookup, is_mandatory}.

    Rules:
      - Skip if Type == 24 (button/navigation) AND IsNavigation == true
      - If IsLookupon == true → lookup field (still include in main fields too)
      - Use DisplayName[0] when non-empty, else fall back to Name

    The field_metadata preserves ALL non-skipped fields (including buttons that
    are NOT navigation) so the pipeline can route them correctly.
    """
    data_fields: List[str] = []
    lookup_fields: List[str] = []
    field_metadata: Dict[str, Dict] = {}

    for f in fields_raw:
        field_type: int = f.get("Type", 0)
        is_nav: bool = f.get("IsNavigation", False)
        is_lookup: bool = f.get("IsLookupon", False)
        is_mandatory: bool = f.get("IsMandatory", False)

        # Skip pure navigation buttons (Type=24 + IsNavigation=true)
        # These represent Selenium click actions with no data value.
        if field_type == NAV_TYPE and is_nav:
            continue

        # Resolve the display name
        display_names: List[str] = f.get("DisplayName", [])
        display_name = display_names[0].strip() if display_names and display_names[0].strip() else ""
        raw_name: str = f.get("Name", "").strip()
        field_name = display_name if display_name else raw_name

        if not field_name:
            continue

        # Store type metadata regardless of lookup status
        field_metadata[field_name] = {
            "type": field_type,
            "is_navigation": is_nav,
            "is_lookup": is_lookup,
            "is_mandatory": is_mandatory,
        }

        if is_lookup:
            lookup_fields.append(field_name)
        else:
            data_fields.append(field_name)

    return data_fields, lookup_fields, field_metadata


# ──────────────────────────────────────────────
# Excel Input Parser (secondary path)
# ──────────────────────────────────────────────

def parse_excel_input(file_path: str | Path) -> Dict[str, Any]:
    """
    Parse an Excel TDM file (.xlsx) into the same normalized structure.

    The Excel layout (TP_AppointmentList_TDM.xlsx):
      - SheetMap:    ExcelSheet → SheetName (TestSuite name mapping)
      - Login Page:  TestsuiteName | ObjectName | Url | UserId | Password | ...
      - Sheet1..N:   Header row with TS name, then:
                     ObjectName | Iteration Value | OCRTextVerification | Field Names | f1 | f2 ...
                     <ObjectName> row with field names in columns 4+
                     Iteration rows: empty col A, iteration number in col B (1, 2, 3...)
      - Lookup:      TestSuiteName | Object | ParentField | Iteration | Field Names | ...

    Args:
        file_path: Path to the .xlsx file

    Returns:
        Normalized template dict (same shape as parse_json_input output)
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel parsing. "
            "Install it with: pip install openpyxl"
        )

    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)

    # ── Step 1: Build SheetMap (ExcelSheet → TestSuite name) ──────────────
    sheet_to_suite: Dict[str, str] = {}  # e.g. {"Sheet1": "TS_01", "Sheet2": "TS_2", ...}
    if "SheetMap" in wb.sheetnames:
        sm_ws = wb["SheetMap"]
        rows = list(sm_ws.iter_rows(values_only=True))
        for row in rows[1:]:  # skip header
            excel_sheet = str(row[0]).strip() if row[0] else ""
            suite_name  = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if excel_sheet and suite_name and suite_name not in ("Login Page", "Lookup"):
                sheet_to_suite[excel_sheet] = suite_name

    package_name = file_path.stem.replace("_TDM", "").replace("TP_", "TP_")

    normalized: Dict[str, Any] = {
        "test_package_name": package_name,
        "test_suites": {},
    }

    # ── Step 2: Parse each data sheet ──────────────────────────────────────
    for excel_sheet_name, suite_name in sheet_to_suite.items():
        if excel_sheet_name not in wb.sheetnames:
            logger.warning(f"[excel_parser] Sheet '{excel_sheet_name}' not found in workbook.")
            continue

        ws = wb[excel_sheet_name]
        objects = _parse_excel_data_sheet(ws, suite_name)

        if objects:
            normalized["test_suites"][suite_name] = {"objects": objects}

    # ── Step 3: If no SheetMap, fall back to scanning all sheets ──────────
    if not normalized["test_suites"]:
        for sheet_name in wb.sheetnames:
            if sheet_name in ("SheetMap", "Login Page", "Lookup"):
                continue
            ws = wb[sheet_name]
            objects = _parse_excel_data_sheet(ws, sheet_name)
            if objects:
                normalized["test_suites"][sheet_name] = {"objects": objects}

    logger.info(
        f"[excel_parser] Parsed Excel: file='{file_path.name}', "
        f"suites={list(normalized['test_suites'].keys())}"
    )
    return normalized


def _parse_excel_data_sheet(ws, suite_name: str) -> Dict[str, Any]:
    """
    Parse a single Excel data sheet (Sheet1, Sheet2, ...) into objects dict.

    Sheet layout:
      Row 0: [TS_XX, '', '', '', '', 'SheetMap', ...]   ← header/title row, skip
      Row 1: ['ObjectName', 'Iteration Value', 'OCRTextVerification', 'Field Names', f1, f2, ...]
      Row 2: ['<actual object name>', '', '', '', <f1_empty>, <f2_empty>, ...]   ← object row
      Row 3+: ['', '1', '', '', ...]   ← iteration row (col B has iteration number)
              ['', '2', '', '', ...]
              ...

    So:
      - Find the header row (col A == 'ObjectName')
      - The next row after header is the object definition row
      - Field names start at column index 3 (0-based) in that row
      - Iteration rows are identified by empty col A + numeric col B
    """
    objects: Dict[str, Any] = {}
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return objects

    # Find the header row
    header_row_idx = -1
    for i, row in enumerate(rows):
        col_a = str(row[0]).strip().lower() if row[0] else ""
        if col_a == "objectname":
            header_row_idx = i
            break

    if header_row_idx < 0:
        # No standard header found — try generic parsing
        return _parse_excel_sheet_generic(ws, suite_name)

    # Parse object + iteration rows after the header
    current_object: Optional[str] = None
    current_fields: List[str] = []
    field_col_start = 3  # fields start at col index 3 (0-based), after ObjectName | Iter | OCR

    for row in rows[header_row_idx + 1:]:
        col_a = str(row[0]).strip() if row[0] is not None else ""
        col_b = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""

        # Check if this is a new object row (col A has a value)
        if col_a:
            # New object definition row
            current_object = col_a
            # Extract field names from column 3 onwards (skip empty ones)
            current_fields = []
            for col_idx in range(field_col_start, len(row)):
                cell_val = row[col_idx]
                if cell_val is not None:
                    field_name = str(cell_val).strip()
                    if field_name:
                        current_fields.append(field_name)

            if current_object and current_object not in objects:
                objects[current_object] = _empty_object_node()
                objects[current_object]["fields"] = list(current_fields)

        elif col_b:
            # Iteration row (col A is empty, col B has iteration number)
            if _is_numeric(col_b) and current_object and current_object in objects:
                iter_num = int(float(col_b))
                obj = objects[current_object]
                obj["iteration_count"] = max(obj["iteration_count"], iter_num)

    return objects


def _parse_excel_sheet_generic(ws, suite_name: str) -> Dict[str, Any]:
    """
    Generic fallback parser: scans column A for non-empty values as field names.
    Used when the standard header structure is not found.
    """
    objects: Dict[str, Any] = {}
    current_object: Optional[str] = None
    iteration_count = 1

    for row in ws.iter_rows(values_only=True):
        if not any(cell is not None for cell in row):
            continue

        col_a = str(row[0]).strip() if row[0] is not None else ""
        col_b = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""

        if not col_a:
            continue

        # Detect object/flow header
        if col_a.lower() in ("object", "flow", "test object", "screen", "module"):
            current_object = col_b if col_b else "Default Object"
            if current_object not in objects:
                objects[current_object] = _empty_object_node()
            continue

        # Detect iteration count
        if col_a.lower() in ("iteration", "iterations", "iteration count", "count"):
            try:
                iteration_count = int(col_b)
                if current_object and current_object in objects:
                    objects[current_object]["iteration_count"] = iteration_count
            except (ValueError, TypeError):
                pass
            continue

        # Detect field rows
        if _looks_like_field_name(col_a):
            if current_object is None:
                current_object = f"{suite_name} - Default"
                objects[current_object] = _empty_object_node()

            if col_a not in objects[current_object]["fields"]:
                objects[current_object]["fields"].append(col_a)

    return objects


def _looks_like_field_name(text: str) -> bool:
    """
    Heuristic: distinguish field names from structural labels/headers.
    Returns True if text looks like a data field name.
    """
    skip_keywords = {
        "field", "fields", "header", "column", "name", "type",
        "mandatory", "lookup", "iteration", "object", "flow",
        "test suite", "test case", "action", "step", "expected",
        "result", "precondition", "description", "approved", "rejected",
        "objectname", "sheetmap",
    }
    lower = text.lower().strip()
    return lower not in skip_keywords and len(text) > 1


def _is_numeric(value: str) -> bool:
    """Return True if value can be interpreted as a number."""
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False


# ──────────────────────────────────────────────
# File-based convenience loaders
# ──────────────────────────────────────────────

def load_json_file(path: str | Path) -> Dict[str, Any]:
    """Load and parse a JSON file, returning the normalized template."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"JSON input file not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return parse_json_input(data)


def load_excel_file(path: str | Path) -> Dict[str, Any]:
    """Load and parse an Excel file, returning the normalized template."""
    return parse_excel_input(path)


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────

def _safe_int(value: Any, default: int = 1) -> int:
    """Safely convert a value to int, returning default on failure."""
    try:
        return int(value)
    except (TypeError, ValueError):
        return default
